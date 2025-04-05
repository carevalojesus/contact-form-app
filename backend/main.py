# main.py
from fastapi import FastAPI, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, EmailStr, Field
from typing import List, Optional, Dict, Any
from datetime import date, datetime, timedelta
import uuid
import json
import os
import io
# Importaciones para manejo de Excel sin pandas directamente
import csv
from openpyxl import load_workbook

# Modelo de datos para el contacto
class ContactBase(BaseModel):
    nombre: str
    apellidos: str
    correo: EmailStr
    celular: str
    fecha_nacimiento: date
    genero: str
    direccion: str

class ContactCreate(ContactBase):
    pass

class Contact(ContactBase):
    id: str
    fecha_registro: datetime

    class Config:
        orm_mode = True

# Función para calcular la edad a partir de la fecha de nacimiento
def calcular_edad(fecha_nacimiento_str: str) -> int:
    try:
        fecha_nacimiento = datetime.strptime(fecha_nacimiento_str, '%Y-%m-%d').date()
        hoy = date.today()
        edad = hoy.year - fecha_nacimiento.year - ((hoy.month, hoy.day) < (fecha_nacimiento.month, fecha_nacimiento.day))
        return edad
    except Exception:
        return 0

# Creación de la aplicación FastAPI
app = FastAPI(title="API de Contactos", 
              description="API para gestionar contactos y exportar a vCard",
              version="1.0.0")

# Configuración de CORS para permitir solicitudes del frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # En producción, especifica los orígenes permitidos
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Simulamos una base de datos con un archivo JSON
DB_FILE = "contacts.json"

# Función para cargar contactos desde el archivo
def load_contacts():
    if not os.path.exists(DB_FILE):
        return []
    
    try:
        with open(DB_FILE, "r", encoding="utf-8") as file:
            return json.load(file)
    except Exception as e:
        print(f"Error al cargar contactos: {e}")
        return []

# Función para guardar contactos en el archivo
def save_contacts(contacts):
    with open(DB_FILE, "w", encoding="utf-8") as file:
        json.dump(contacts, file, default=str, ensure_ascii=False, indent=2)

# Rutas para la API
@app.get("/")
def read_root():
    return {"message": "API de Contactos funcionando correctamente"}

@app.get("/contacts/", response_model=List[Dict[str, Any]])
def get_contacts():
    contacts = load_contacts()
    # Añadir la edad a cada contacto
    for contact in contacts:
        contact["edad"] = calcular_edad(contact["fecha_nacimiento"])
    return contacts

@app.post("/contacts/", response_model=Contact, status_code=201)
def create_contact(contact: ContactCreate):
    contacts = load_contacts()
    
    # Creamos un nuevo contacto con ID y fecha de registro
    new_contact = {
        **contact.dict(),
        "id": str(uuid.uuid4()),
        "fecha_registro": datetime.now()
    }
    
    contacts.append(new_contact)
    save_contacts(contacts)
    return new_contact

@app.get("/contacts/{contact_id}")
def get_contact(contact_id: str):
    contacts = load_contacts()
    for contact in contacts:
        if contact["id"] == contact_id:
            # Añadir la edad al contacto
            contact["edad"] = calcular_edad(contact["fecha_nacimiento"])
            return contact
    raise HTTPException(status_code=404, detail="Contacto no encontrado")

@app.put("/contacts/{contact_id}", response_model=Contact)
def update_contact(contact_id: str, contact_update: ContactCreate):
    contacts = load_contacts()
    
    for i, contact in enumerate(contacts):
        if contact["id"] == contact_id:
            # Actualizar los campos del contacto pero mantener id y fecha_registro
            updated_contact = {
                **contact_update.dict(),
                "id": contact["id"],
                "fecha_registro": contact["fecha_registro"]
            }
            contacts[i] = updated_contact
            save_contacts(contacts)
            return updated_contact
    
    raise HTTPException(status_code=404, detail="Contacto no encontrado")

@app.delete("/contacts/{contact_id}")
def delete_contact(contact_id: str):
    contacts = load_contacts()
    initial_length = len(contacts)
    
    contacts = [contact for contact in contacts if contact["id"] != contact_id]
    
    if len(contacts) == initial_length:
        raise HTTPException(status_code=404, detail="Contacto no encontrado")
    
    save_contacts(contacts)
    return {"message": "Contacto eliminado correctamente"}

@app.get("/contacts/{contact_id}/vcard")
def get_contact_vcard(contact_id: str):
    contacts = load_contacts()
    
    for contact in contacts:
        if contact["id"] == contact_id:
            # Crear vCard según el estándar
            vcard = f"""BEGIN:VCARD
VERSION:3.0
N:{contact['apellidos']};{contact['nombre']};;;
FN:{contact['nombre']} {contact['apellidos']}
TEL;TYPE=CELL:{contact['celular']}
EMAIL:{contact['correo']}
BDAY:{contact['fecha_nacimiento']}
GENDER:{contact['genero']}
ADR:;;{contact['direccion']};;;
END:VCARD
"""
            return {"vcard": vcard}
    
    raise HTTPException(status_code=404, detail="Contacto no encontrado")

@app.get("/export/vcard")
def export_all_vcards():
    contacts = load_contacts()
    
    if not contacts:
        raise HTTPException(status_code=404, detail="No hay contactos para exportar")
    
    all_vcards = ""
    
    for contact in contacts:
        vcard = f"""BEGIN:VCARD
VERSION:3.0
N:{contact['apellidos']};{contact['nombre']};;;
FN:{contact['nombre']} {contact['apellidos']}
TEL;TYPE=CELL:{contact['celular']}
EMAIL:{contact['correo']}
BDAY:{contact['fecha_nacimiento']}
GENDER:{contact['genero']}
ADR:;;{contact['direccion']};;;
END:VCARD
"""
        all_vcards += vcard + "\n"
    
    return {"vcard": all_vcards}

@app.post("/import/excel")
async def import_from_excel(file: UploadFile = File(...)):
    if not file.filename.endswith(('.xls', '.xlsx', '.csv')):
        raise HTTPException(status_code=400, detail="El archivo debe ser Excel (.xls, .xlsx) o CSV (.csv)")
    
    try:
        print(f"Procesando archivo: {file.filename}")
        # Leer el contenido del archivo
        contents = await file.read()
        contents_io = io.BytesIO(contents)
        
        # Lista para almacenar los datos
        data = []
        headers = []
        
        # Comprobar si es un archivo CSV
        if file.filename.endswith('.csv'):
            print("Procesando archivo CSV")
            try:
                # Decodificar el contenido en texto - intentar diferentes encodings
                for encoding in ['utf-8-sig', 'utf-8', 'latin-1', 'iso-8859-1']:
                    try:
                        text_content = contents.decode(encoding)
                        csv_reader = csv.reader(text_content.splitlines())
                        headers = next(csv_reader, [])  # Primera fila como encabezados
                        print(f"Encabezados encontrados: {headers}")
                        
                        for row in csv_reader:
                            if len(row) == len(headers):
                                # Mapear cada valor a su columna
                                row_dict = {headers[i]: row[i] for i in range(len(headers))}
                                data.append(row_dict)
                        break
                    except UnicodeDecodeError:
                        continue
            except Exception as csv_error:
                print(f"Error procesando CSV: {str(csv_error)}")
                raise
                
        # Si es un archivo Excel
        else:
            print("Procesando archivo Excel")
            try:
                # Usar openpyxl para leer el Excel
                workbook = load_workbook(contents_io)
                worksheet = workbook.active
                
                # Obtener encabezados (primera fila)
                for cell in worksheet[1]:
                    if cell.value is not None:
                        headers.append(cell.value.strip() if isinstance(cell.value, str) else cell.value)
                
                print(f"Encabezados encontrados: {headers}")
                
                # Procesar filas de datos
                for row in worksheet.iter_rows(min_row=2):
                    row_dict = {}
                    for i, cell in enumerate(row):
                        if i < len(headers) and headers[i] is not None:
                            row_dict[headers[i]] = cell.value
                    data.append(row_dict)
            except Exception as excel_error:
                print(f"Error procesando Excel: {str(excel_error)}")
                raise
        
        print(f"Datos extraídos: {len(data)} filas")
        
        # Comprobar que las columnas necesarias estén presentes
        required_columns = ['nombre', 'apellidos', 'correo', 'celular', 'fecha_nacimiento', 'genero', 'direccion']
        lowercase_headers = [h.lower() if isinstance(h, str) else h for h in headers]
        
        # Mapeamos los encabezados para manejar variaciones en mayúsculas/minúsculas
        header_mapping = {}
        for i, header in enumerate(lowercase_headers):
            if isinstance(header, str):
                for req in required_columns:
                    if req.lower() == header.lower():
                        header_mapping[req] = headers[i]
        
        print(f"Mapeo de encabezados: {header_mapping}")
        
        missing_columns = [col for col in required_columns if col not in header_mapping]
        
        if missing_columns:
            missing_msg = f"El archivo no contiene las columnas requeridas: {', '.join(missing_columns)}"
            print(missing_msg)
            raise HTTPException(status_code=400, detail=missing_msg)
        
        # Cargar contactos existentes
        contacts = load_contacts()
        
        # Contador de contactos añadidos
        added_contacts = 0
        
        # Procesar cada fila de datos
        for row in data:
            try:
                # Crear diccionario con los nombres de columna estandarizados
                std_row = {}
                for req_col in required_columns:
                    mapped_col = header_mapping.get(req_col)
                    if mapped_col in row and row[mapped_col] is not None:
                        std_row[req_col] = row[mapped_col]
                    else:
                        std_row[req_col] = ''  # Valor por defecto
                
                # Validar datos requeridos - todos deben tener al menos nombre o apellido
                if not std_row['nombre'] and not std_row['apellidos']:
                    print(f"Fila saltada porque nombre y apellidos están vacíos: {std_row}")
                    continue
                
                # Convertir fecha al formato correcto
                fecha_str = str(std_row['fecha_nacimiento'] or '')
                fecha_nacimiento = fecha_str
                
                # Intentar formatear diferentes posibles formatos de fecha
                try:
                    # Si es datetime de Excel (número flotante)
                    if isinstance(std_row['fecha_nacimiento'], (int, float)):
                        from datetime import timedelta
                        # Excel guarda fechas como días desde 1900-01-01
                        fecha_base = date(1899, 12, 30)  
                        fecha_obj = fecha_base + timedelta(days=int(std_row['fecha_nacimiento']))
                        fecha_nacimiento = fecha_obj.isoformat()
                    # Si es string en formato fecha
                    elif isinstance(fecha_str, str) and fecha_str.strip():
                        # Intentar varios formatos comunes
                        for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y']:
                            try:
                                fecha_obj = datetime.strptime(fecha_str, fmt).date()
                                fecha_nacimiento = fecha_obj.isoformat()
                                break
                            except ValueError:
                                continue
                except Exception as date_error:
                    print(f"Error procesando fecha '{fecha_str}': {str(date_error)}")
                    # Si falla, usar fecha actual
                    fecha_nacimiento = date.today().isoformat()
                
                # Asegurar valores string válidos para todos los campos
                for key in std_row:
                    if std_row[key] is None:
                        std_row[key] = ''
                    elif not isinstance(std_row[key], str):
                        std_row[key] = str(std_row[key])
                
                # Crear un nuevo contacto
                new_contact = {
                    "nombre": std_row['nombre'],
                    "apellidos": std_row['apellidos'],
                    "correo": std_row['correo'],
                    "celular": std_row['celular'],
                    "fecha_nacimiento": fecha_nacimiento,
                    "genero": std_row['genero'],
                    "direccion": std_row['direccion'],
                    "id": str(uuid.uuid4()),
                    "fecha_registro": datetime.now().isoformat()
                }
                
                contacts.append(new_contact)
                added_contacts += 1
                
            except Exception as row_error:
                print(f"Error procesando fila: {str(row_error)}")
                # Continuar con la siguiente fila en caso de error
                continue
        
        # Guardar todos los contactos
        save_contacts(contacts)
        
        success_msg = f"Se importaron {added_contacts} contactos correctamente"
        print(success_msg)
        return {"message": success_msg}
    
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        error_msg = f"Error al procesar el archivo: {str(e)}"
        print(error_msg)
        print(error_trace)
        raise HTTPException(status_code=500, detail=error_msg)

# Ejecutar la aplicación con Uvicorn
if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)